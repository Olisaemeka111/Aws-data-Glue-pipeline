#!/usr/bin/env python3
"""
Convert ACTUALLY_DEPLOYED_RESOURCES.csv to Excel with charts and visual analysis
Enhanced version with comprehensive data visualization
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, BarChart, LineChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
import sys
import os

def create_excel_with_charts(csv_file, output_file):
    """Convert CSV to Excel with charts and visual analysis"""
    
    print(f"🔄 Reading CSV file: {csv_file}")
    
    # Read the CSV file
    try:
        df = pd.read_csv(csv_file)
        print(f"✅ Successfully read {len(df)} rows")
    except Exception as e:
        print(f"❌ Error reading CSV: {e}")
        return False
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create dashboard sheet with charts
    create_dashboard_sheet(wb, df)
    
    # Create summary sheet with enhanced visuals
    create_enhanced_summary_sheet(wb, df)
    
    # Create module analysis sheet with charts
    create_module_analysis_sheet(wb, df)
    
    # Create detailed sheet with grouping
    create_detailed_sheet(wb, df)
    
    # Create module-specific sheets
    create_module_sheets(wb, df)
    
    # Save the workbook
    try:
        wb.save(output_file)
        print(f"✅ Excel file with charts saved: {output_file}")
        return True
    except Exception as e:
        print(f"❌ Error saving Excel: {e}")
        return False

def create_dashboard_sheet(wb, df):
    """Create executive dashboard with charts and KPIs"""
    
    ws = wb.create_sheet("📊 Dashboard")
    
    # Title
    ws['A1'] = "AWS Glue ETL Pipeline - Executive Dashboard"
    title_cell = ws['A1']
    title_cell.font = Font(size=20, bold=True, color="1F4E79")
    title_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    ws.merge_cells('A1:L1')
    title_cell.alignment = Alignment(horizontal='center')
    
    # Calculate statistics
    total_resources = len(df)
    deployed_resources = len(df[df['Status'] == '✅ Deployed'])
    missing_resources = len(df[df['Status'] == '❌ Missing'])
    success_rate = (deployed_resources / total_resources * 100)
    
    # KPI Cards
    kpi_data = [
        ("Total Resources", total_resources, "2F5597"),
        ("Deployed", deployed_resources, "28A745"),
        ("Missing", missing_resources, "DC3545"),
        ("Success Rate", f"{success_rate:.1f}%", "17A2B8")
    ]
    
    # Create KPI cards
    for i, (label, value, color) in enumerate(kpi_data):
        col = i * 3 + 1
        
        # Label
        label_cell = ws.cell(row=3, column=col, value=label)
        label_cell.font = Font(size=12, bold=True, color="FFFFFF")
        label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        label_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(f'{openpyxl.utils.get_column_letter(col)}3:{openpyxl.utils.get_column_letter(col+1)}3')
        
        # Value
        value_cell = ws.cell(row=4, column=col, value=value)
        value_cell.font = Font(size=16, bold=True, color=color)
        value_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(f'{openpyxl.utils.get_column_letter(col)}4:{openpyxl.utils.get_column_letter(col+1)}4')
    
    # Prepare data for charts
    module_stats = df.groupby('Module').agg({
        'Resource Name': 'count',
        'Status': lambda x: (x == '✅ Deployed').sum()
    }).rename(columns={'Resource Name': 'Total', 'Status': 'Deployed'})
    
    module_stats['Missing'] = module_stats['Total'] - module_stats['Deployed']
    module_stats['Success Rate'] = (module_stats['Deployed'] / module_stats['Total'] * 100).round(1)
    
    # Add data for charts starting at row 6
    ws['A6'] = "Module"
    ws['B6'] = "Total"
    ws['C6'] = "Deployed"
    ws['D6'] = "Missing"
    ws['E6'] = "Success Rate"
    
    # Format headers
    for col in range(1, 6):
        cell = ws.cell(row=6, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add module data
    for i, (module, stats) in enumerate(module_stats.iterrows()):
        row = 7 + i
        ws.cell(row=row, column=1, value=module)
        ws.cell(row=row, column=2, value=stats['Total'])
        ws.cell(row=row, column=3, value=stats['Deployed'])
        ws.cell(row=row, column=4, value=stats['Missing'])
        ws.cell(row=row, column=5, value=stats['Success Rate'])
    
    # Create Overall Status Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "Overall Deployment Status"
    pie_chart.title.tx.rich.p[0].pPr = ParagraphProperties()
    pie_chart.title.tx.rich.p[0].pPr.defRPr = CharacterProperties(sz=1400, b=True)
    
    # Add data for pie chart
    ws['H6'] = "Status"
    ws['I6'] = "Count"
    ws['H7'] = "Deployed"
    ws['I7'] = deployed_resources
    ws['H8'] = "Missing"
    ws['I8'] = missing_resources
    
    # Format pie chart data
    for row in range(6, 9):
        for col in range(8, 10):
            cell = ws.cell(row=row, column=col)
            if row == 6:
                cell.font = Font(bold=True)
    
    pie_data = Reference(ws, min_col=9, min_row=7, max_row=8)
    pie_labels = Reference(ws, min_col=8, min_row=7, max_row=8)
    pie_chart.add_data(pie_data)
    pie_chart.set_categories(pie_labels)
    
    # Chart styling will use default colors
    
    # Position pie chart
    pie_chart.width = 12
    pie_chart.height = 8
    ws.add_chart(pie_chart, "H10")
    
    # Create Module Performance Bar Chart
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Deployment Status by Module"
    bar_chart.title.tx.rich.p[0].pPr = ParagraphProperties()
    bar_chart.title.tx.rich.p[0].pPr.defRPr = CharacterProperties(sz=1400, b=True)
    
    # Data for bar chart
    bar_data = Reference(ws, min_col=3, min_row=6, max_col=4, max_row=6 + len(module_stats))
    bar_categories = Reference(ws, min_col=1, min_row=7, max_row=6 + len(module_stats))
    
    bar_chart.add_data(bar_data, titles_from_data=True)
    bar_chart.set_categories(bar_categories)
    
    # Chart styling will use default colors
    
    bar_chart.x_axis.title = "Modules"
    bar_chart.y_axis.title = "Number of Resources"
    bar_chart.width = 12
    bar_chart.height = 8
    
    ws.add_chart(bar_chart, "A10")
    
    # Success Rate Line Chart
    line_chart = LineChart()
    line_chart.title = "Success Rate by Module"
    line_chart.title.tx.rich.p[0].pPr = ParagraphProperties()
    line_chart.title.tx.rich.p[0].pPr.defRPr = CharacterProperties(sz=1400, b=True)
    
    line_data = Reference(ws, min_col=5, min_row=6, max_row=6 + len(module_stats))
    line_categories = Reference(ws, min_col=1, min_row=7, max_row=6 + len(module_stats))
    
    line_chart.add_data(line_data, titles_from_data=True)
    line_chart.set_categories(line_categories)
    
    line_chart.y_axis.title = "Success Rate (%)"
    line_chart.x_axis.title = "Modules"
    line_chart.y_axis.scaling.min = 0
    line_chart.y_axis.scaling.max = 100
    line_chart.width = 12
    line_chart.height = 8
    
    ws.add_chart(line_chart, "A25")
    
    # Auto-fit columns
    for col_num in range(1, 13):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

def create_enhanced_summary_sheet(wb, df):
    """Create enhanced summary sheet with additional charts"""
    
    ws = wb.create_sheet("📋 Summary")
    
    # Title
    ws['A1'] = "AWS Glue ETL Pipeline - Deployment Summary"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws.merge_cells('A1:E1')
    
    # Overall statistics
    total_resources = len(df)
    deployed_resources = len(df[df['Status'] == '✅ Deployed'])
    missing_resources = len(df[df['Status'] == '❌ Missing'])
    
    # Summary table
    summary_data = [
        ["Metric", "Count", "Percentage"],
        ["Total Resources", total_resources, "100%"],
        ["Successfully Deployed", deployed_resources, f"{deployed_resources/total_resources*100:.1f}%"],
        ["Missing/Failed", missing_resources, f"{missing_resources/total_resources*100:.1f}%"],
    ]
    
    start_row = 3
    for i, row in enumerate(summary_data):
        for j, value in enumerate(row):
            cell = ws.cell(row=start_row + i, column=j + 1, value=value)
            if i == 0:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            elif j == 2 and i > 0:  # Percentage column
                if "100%" in str(value):
                    cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
                elif missing_resources > 0 and "Missing" in summary_data[i][0]:
                    cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
    
    # Module breakdown
    module_stats = df.groupby('Module').agg({
        'Resource Name': 'count',
        'Status': lambda x: (x == '✅ Deployed').sum()
    }).rename(columns={'Resource Name': 'Total', 'Status': 'Deployed'})
    
    module_stats['Missing'] = module_stats['Total'] - module_stats['Deployed']
    module_stats['Success Rate'] = (module_stats['Deployed'] / module_stats['Total'] * 100).round(1)
    
    # Module breakdown table
    ws['A8'] = "Deployment Status by Module"
    ws['A8'].font = Font(size=14, bold=True, color="1F4E79")
    
    module_headers = ["Module", "Total Resources", "Deployed", "Missing", "Success Rate (%)"]
    for j, header in enumerate(module_headers):
        cell = ws.cell(row=10, column=j + 1, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    for i, (module, stats) in enumerate(module_stats.iterrows()):
        row_num = 11 + i
        ws.cell(row=row_num, column=1, value=module)
        ws.cell(row=row_num, column=2, value=stats['Total'])
        ws.cell(row=row_num, column=3, value=stats['Deployed'])
        ws.cell(row=row_num, column=4, value=stats['Missing'])
        
        success_rate_cell = ws.cell(row=row_num, column=5, value=stats['Success Rate'])
        
        # Color code success rate
        if stats['Success Rate'] == 100:
            success_rate_cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
        elif stats['Success Rate'] < 100:
            success_rate_cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
    
    # Create Resource Type Distribution Chart
    resource_types = df['Resource Type'].value_counts().head(10)
    
    # Add resource type data for chart
    ws['G8'] = "Resource Type Distribution (Top 10)"
    ws['G8'].font = Font(size=12, bold=True, color="1F4E79")
    
    ws['G10'] = "Resource Type"
    ws['H10'] = "Count"
    
    for i, (resource_type, count) in enumerate(resource_types.items()):
        row = 11 + i
        ws.cell(row=row, column=7, value=resource_type)
        ws.cell(row=row, column=8, value=count)
    
    # Create horizontal bar chart for resource types
    resource_chart = BarChart()
    resource_chart.type = "bar"
    resource_chart.title = "Resource Type Distribution"
    
    resource_data = Reference(ws, min_col=8, min_row=11, max_row=10 + len(resource_types))
    resource_categories = Reference(ws, min_col=7, min_row=11, max_row=10 + len(resource_types))
    
    resource_chart.add_data(resource_data)
    resource_chart.set_categories(resource_categories)
    resource_chart.width = 12
    resource_chart.height = 10
    
    ws.add_chart(resource_chart, "G22")
    
    # Auto-fit columns
    for col_num in range(1, 9):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_module_analysis_sheet(wb, df):
    """Create detailed module analysis with charts"""
    
    ws = wb.create_sheet("📊 Module Analysis")
    
    # Title
    ws['A1'] = "Module-wise Deployment Analysis"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws.merge_cells('A1:H1')
    
    # Calculate module statistics
    module_stats = df.groupby('Module').agg({
        'Resource Name': 'count',
        'Status': lambda x: (x == '✅ Deployed').sum()
    }).rename(columns={'Resource Name': 'Total', 'Status': 'Deployed'})
    
    module_stats['Missing'] = module_stats['Total'] - module_stats['Deployed']
    module_stats['Success Rate'] = (module_stats['Deployed'] / module_stats['Total'] * 100).round(1)
    
    # Add detailed module table
    headers = ["Module", "Total", "Deployed", "Missing", "Success Rate (%)", "Status"]
    for j, header in enumerate(headers):
        cell = ws.cell(row=3, column=j + 1, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    for i, (module, stats) in enumerate(module_stats.iterrows()):
        row_num = 4 + i
        ws.cell(row=row_num, column=1, value=module)
        ws.cell(row=row_num, column=2, value=stats['Total'])
        ws.cell(row=row_num, column=3, value=stats['Deployed'])
        ws.cell(row=row_num, column=4, value=stats['Missing'])
        ws.cell(row=row_num, column=5, value=stats['Success Rate'])
        
        # Status indicator
        status = "✅ Complete" if stats['Success Rate'] == 100 else "⚠️ Issues"
        status_cell = ws.cell(row=row_num, column=6, value=status)
        
        if stats['Success Rate'] == 100:
            status_cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
    
    # Create stacked bar chart
    stacked_chart = BarChart()
    stacked_chart.type = "col"
    stacked_chart.grouping = "stacked"
    stacked_chart.title = "Module Resource Distribution (Deployed vs Missing)"
    
    deployed_data = Reference(ws, min_col=3, min_row=3, max_row=3 + len(module_stats))
    missing_data = Reference(ws, min_col=4, min_row=3, max_row=3 + len(module_stats))
    categories = Reference(ws, min_col=1, min_row=4, max_row=3 + len(module_stats))
    
    stacked_chart.add_data(deployed_data, titles_from_data=True)
    stacked_chart.add_data(missing_data, titles_from_data=True)
    stacked_chart.set_categories(categories)
    
    # Customize colors
    stacked_chart.series[0].graphicalProperties.solidFill = "28A745"  # Green
    stacked_chart.series[1].graphicalProperties.solidFill = "DC3545"  # Red
    
    stacked_chart.width = 15
    stacked_chart.height = 10
    ws.add_chart(stacked_chart, "A12")
    
    # Create success rate gauge chart (using line chart)
    gauge_chart = LineChart()
    gauge_chart.title = "Success Rate Trends"
    
    success_data = Reference(ws, min_col=5, min_row=3, max_row=3 + len(module_stats))
    gauge_categories = Reference(ws, min_col=1, min_row=4, max_row=3 + len(module_stats))
    
    gauge_chart.add_data(success_data, titles_from_data=True)
    gauge_chart.set_categories(gauge_categories)
    gauge_chart.y_axis.scaling.min = 0
    gauge_chart.y_axis.scaling.max = 100
    gauge_chart.width = 15
    gauge_chart.height = 8
    
    ws.add_chart(gauge_chart, "A27")
    
    # Auto-fit columns
    for col_num in range(1, 7):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 18

def create_detailed_sheet(wb, df):
    """Create detailed sheet with all resources grouped by module"""
    
    ws = wb.create_sheet("📝 All Resources")
    
    # Title
    ws['A1'] = "AWS Glue ETL Pipeline - Detailed Resource List"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws.merge_cells('A1:E1')
    
    # Headers
    headers = ["Module", "Resource Type", "Resource Name", "Description", "Status"]
    for j, header in enumerate(headers):
        cell = ws.cell(row=3, column=j + 1, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Group data by module and add rows
    current_row = 4
    for module in df['Module'].unique():
        module_data = df[df['Module'] == module].copy()
        
        # Add module separator row
        module_cell = ws.cell(row=current_row, column=1, value=f"📁 {module.upper()}")
        module_cell.font = Font(bold=True, size=12, color="1F4E79")
        module_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        # Add module resources
        for _, row in module_data.iterrows():
            for j, value in enumerate(row):
                cell = ws.cell(row=current_row, column=j + 1, value=value)
                
                # Format status column
                if j == 4:  # Status column
                    if '✅' in str(value):
                        cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
                    elif '❌' in str(value):
                        cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
                
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            current_row += 1
        
        # Add space between modules
        current_row += 1
    
    # Auto-fit columns
    for col_num in range(1, 6):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_module_sheets(wb, df):
    """Create individual sheets for each module"""
    
    for module in df['Module'].unique():
        module_data = df[df['Module'] == module].copy()
        
        # Create sheet with clean name
        sheet_name = module.replace('_', ' ').title()[:31]
        ws = wb.create_sheet(f"🗂️ {sheet_name}")
        
        # Title
        ws['A1'] = f"Module: {module.upper()}"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
        ws.merge_cells('A1:E1')
        
        # Statistics
        total = len(module_data)
        deployed = len(module_data[module_data['Status'] == '✅ Deployed'])
        missing = total - deployed
        success_rate = (deployed / total * 100) if total > 0 else 0
        
        # KPI row
        ws['A3'] = f"Total: {total}"
        ws['B3'] = f"Deployed: {deployed}"
        ws['C3'] = f"Missing: {missing}"
        ws['D3'] = f"Success: {success_rate:.1f}%"
        
        # Color code KPIs
        ws['B3'].fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
        if missing > 0:
            ws['C3'].fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
        
        # Headers
        headers = ["Resource Type", "Resource Name", "Description", "Status"]
        for j, header in enumerate(headers):
            cell = ws.cell(row=5, column=j + 1, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for i, (_, row) in enumerate(module_data.iterrows()):
            row_num = 6 + i
            row_data = [row['Resource Type'], row['Resource Name'], row['Description'], row['Status']]
            
            for j, value in enumerate(row_data):
                cell = ws.cell(row=row_num, column=j + 1, value=value)
                
                # Format status column
                if j == 3:  # Status column
                    if '✅' in str(value):
                        cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
                    elif '❌' in str(value):
                        cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
                
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-fit columns
        for col_num in range(1, 5):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_num)
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

def main():
    """Main function"""
    
    csv_file = "ACTUALLY_DEPLOYED_RESOURCES.csv"
    output_file = "AWS_Glue_Pipeline_Resources_with_Charts.xlsx"
    
    print("📊 Converting CSV to Excel with comprehensive charts and analysis...")
    print(f"📁 Input: {csv_file}")
    print(f"📁 Output: {output_file}")
    print()
    
    if not os.path.exists(csv_file):
        print(f"❌ Error: CSV file '{csv_file}' not found")
        return False
    
    success = create_excel_with_charts(csv_file, output_file)
    
    if success:
        print()
        print("✅ Enhanced Excel conversion completed successfully!")
        print(f"📊 Excel file with charts created: {output_file}")
        print()
        print("📋 Sheets created:")
        print("   • 📊 Dashboard - Executive dashboard with KPIs and charts")
        print("   • 📋 Summary - Enhanced summary with resource type distribution")
        print("   • 📊 Module Analysis - Detailed module performance with charts")
        print("   • 📝 All Resources - Complete list with visual grouping")
        print("   • 🗂️ Individual module sheets")
        print()
        print("📈 Charts included:")
        print("   • Overall deployment status pie chart")
        print("   • Module performance bar charts")
        print("   • Success rate trend lines")
        print("   • Resource type distribution charts")
        print("   • Stacked deployment status charts")
        print()
        print("💡 Enhanced features:")
        print("   • Executive dashboard with KPIs")
        print("   • Color-coded status indicators")
        print("   • Multiple chart types for different perspectives")
        print("   • Professional formatting and layout")
        print("   • Interactive visual analysis")
        
        return True
    else:
        print("❌ Conversion failed!")
        return False

if __name__ == "__main__":
    main() 