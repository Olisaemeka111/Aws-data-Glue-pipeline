#!/usr/bin/env python3
"""
Convert ACTUALLY_DEPLOYED_RESOURCES.csv to Excel with grouping and formatting
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys
import os

def create_excel_with_grouping(csv_file, output_file):
    """Convert CSV to Excel with grouping and formatting"""
    
    print(f"🔄 Reading CSV file: {csv_file}")
    
    # Read the CSV file
    try:
        df = pd.read_csv(csv_file)
        print(f"✅ Successfully read {len(df)} rows")
    except Exception as e:
        print(f"❌ Error reading CSV: {e}")
        return False
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create summary sheet
    create_summary_sheet(wb, df)
    
    # Create detailed sheet with grouping
    create_detailed_sheet(wb, df)
    
    # Create module-specific sheets
    create_module_sheets(wb, df)
    
    # Save the workbook
    try:
        wb.save(output_file)
        print(f"✅ Excel file saved: {output_file}")
        return True
    except Exception as e:
        print(f"❌ Error saving Excel: {e}")
        return False

def create_summary_sheet(wb, df):
    """Create summary sheet with overall statistics"""
    
    ws = wb.create_sheet("Summary")
    
    # Title
    ws['A1'] = "AWS Glue ETL Pipeline - Resource Deployment Summary"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws.merge_cells('A1:E1')
    
    # Overall statistics
    total_resources = len(df)
    deployed_resources = len(df[df['Status'] == '✅ Deployed'])
    missing_resources = len(df[df['Status'] == '❌ Missing'])
    
    # Summary table
    summary_data = [
        ["Metric", "Count", "Percentage"],
        ["Total Resources", total_resources, "100%"],
        ["Successfully Deployed", deployed_resources, f"{deployed_resources/total_resources*100:.1f}%"],
        ["Missing/Failed", missing_resources, f"{missing_resources/total_resources*100:.1f}%"],
    ]
    
    start_row = 3
    for i, row in enumerate(summary_data):
        for j, value in enumerate(row):
            cell = ws.cell(row=start_row + i, column=j + 1, value=value)
            if i == 0:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            elif j == 2 and i > 0:  # Percentage column
                if "100%" in str(value):
                    cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
                elif missing_resources > 0 and "Missing" in summary_data[i][0]:
                    cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
    
    # Module breakdown
    module_stats = df.groupby('Module').agg({
        'Resource Name': 'count',
        'Status': lambda x: (x == '✅ Deployed').sum()
    }).rename(columns={'Resource Name': 'Total', 'Status': 'Deployed'})
    
    module_stats['Missing'] = module_stats['Total'] - module_stats['Deployed']
    module_stats['Success Rate'] = (module_stats['Deployed'] / module_stats['Total'] * 100).round(1)
    
    # Module breakdown table
    ws['A8'] = "Deployment Status by Module"
    ws['A8'].font = Font(size=14, bold=True, color="1F4E79")
    
    module_headers = ["Module", "Total Resources", "Deployed", "Missing", "Success Rate (%)"]
    for j, header in enumerate(module_headers):
        cell = ws.cell(row=10, column=j + 1, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    for i, (module, stats) in enumerate(module_stats.iterrows()):
        row_num = 11 + i
        ws.cell(row=row_num, column=1, value=module)
        ws.cell(row=row_num, column=2, value=stats['Total'])
        ws.cell(row=row_num, column=3, value=stats['Deployed'])
        ws.cell(row=row_num, column=4, value=stats['Missing'])
        
        success_rate_cell = ws.cell(row=row_num, column=5, value=stats['Success Rate'])
        
        # Color code success rate
        if stats['Success Rate'] == 100:
            success_rate_cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
        elif stats['Success Rate'] < 100:
            success_rate_cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
    
    # Auto-fit columns
    for col_num in range(1, 6):  # A to E columns
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_detailed_sheet(wb, df):
    """Create detailed sheet with all resources grouped by module"""
    
    ws = wb.create_sheet("All Resources")
    
    # Title
    ws['A1'] = "AWS Glue ETL Pipeline - Detailed Resource List"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws.merge_cells('A1:E1')
    
    # Headers
    headers = ["Module", "Resource Type", "Resource Name", "Description", "Status"]
    for j, header in enumerate(headers):
        cell = ws.cell(row=3, column=j + 1, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Group data by module and add rows
    current_row = 4
    for module in df['Module'].unique():
        module_data = df[df['Module'] == module].copy()
        
        # Add module separator row
        module_cell = ws.cell(row=current_row, column=1, value=f"📁 {module.upper()}")
        module_cell.font = Font(bold=True, size=12, color="1F4E79")
        module_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        # Add module resources
        for _, row in module_data.iterrows():
            for j, value in enumerate(row):
                cell = ws.cell(row=current_row, column=j + 1, value=value)
                
                # Format status column
                if j == 4:  # Status column
                    if '✅' in str(value):
                        cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
                    elif '❌' in str(value):
                        cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
                
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            current_row += 1
        
        # Add space between modules
        current_row += 1
    
    # Auto-fit columns
    for col_num in range(1, 6):  # A to E columns
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_module_sheets(wb, df):
    """Create individual sheets for each module"""
    
    for module in df['Module'].unique():
        module_data = df[df['Module'] == module].copy()
        
        # Create sheet with clean name
        sheet_name = module.replace('_', ' ').title()[:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)
        
        # Title
        ws['A1'] = f"Module: {module.upper()}"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
        ws.merge_cells('A1:E1')
        
        # Statistics
        total = len(module_data)
        deployed = len(module_data[module_data['Status'] == '✅ Deployed'])
        missing = total - deployed
        success_rate = (deployed / total * 100) if total > 0 else 0
        
        ws['A3'] = f"Total Resources: {total}"
        ws['A4'] = f"Deployed: {deployed}"
        ws['A5'] = f"Missing: {missing}"
        ws['A6'] = f"Success Rate: {success_rate:.1f}%"
        
        # Headers
        headers = ["Resource Type", "Resource Name", "Description", "Status"]
        for j, header in enumerate(headers):
            cell = ws.cell(row=8, column=j + 1, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data (excluding Module column)
        for i, (_, row) in enumerate(module_data.iterrows()):
            row_num = 9 + i
            row_data = [row['Resource Type'], row['Resource Name'], row['Description'], row['Status']]
            
            for j, value in enumerate(row_data):
                cell = ws.cell(row=row_num, column=j + 1, value=value)
                
                # Format status column
                if j == 3:  # Status column
                    if '✅' in str(value):
                        cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
                    elif '❌' in str(value):
                        cell.fill = PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")
                
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-fit columns
        for col_num in range(1, 5):  # A to D columns for module sheets
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_num)
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

def main():
    """Main function"""
    
    csv_file = "ACTUALLY_DEPLOYED_RESOURCES.csv"
    output_file = "AWS_Glue_Pipeline_Resources.xlsx"
    
    print("🔄 Converting CSV to Excel with grouping...")
    print(f"📁 Input: {csv_file}")
    print(f"📁 Output: {output_file}")
    print()
    
    if not os.path.exists(csv_file):
        print(f"❌ Error: CSV file '{csv_file}' not found")
        return False
    
    success = create_excel_with_grouping(csv_file, output_file)
    
    if success:
        print()
        print("✅ Conversion completed successfully!")
        print(f"📊 Excel file created: {output_file}")
        print()
        print("📋 Sheets created:")
        print("   • Summary - Overall statistics and module breakdown")
        print("   • All Resources - Complete list with grouping") 
        print("   • Individual module sheets for detailed view")
        print()
        print("💡 Features included:")
        print("   • Color-coded status (Green=Deployed, Red=Missing)")
        print("   • Module grouping and statistics")
        print("   • Auto-fitted columns")
        print("   • Professional formatting")
        
        return True
    else:
        print("❌ Conversion failed!")
        return False

if __name__ == "__main__":
    main() 